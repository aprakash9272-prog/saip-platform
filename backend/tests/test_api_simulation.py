from io import BytesIO

from openpyxl import load_workbook


def _build_scenario(client):
    domain = client.post("/domains", json={"name": "Endpoint Security"}).json()
    capability = client.post(
        "/capabilities", json={"name": "EDR", "code": "SIMAPI-001", "domain_id": domain["id"]}
    ).json()

    vendor = client.post("/vendors", json={"name": "Acme"}).json()
    product = client.post(
        "/products", json={"name": "Shield", "vendor_id": vendor["id"]}
    ).json()
    edition = client.post(
        "/editions", json={"name": "Pro", "product_id": product["id"]}
    ).json()
    module = client.post(
        "/modules",
        json={"name": "Detector", "edition_id": edition["id"], "capability_ids": [capability["id"]]},
    ).json()

    customer = client.post("/customers", json={"name": "Acme Corp"}).json()
    environment = client.post(
        "/environments",
        json={"name": "Production", "environment_type": "Production", "customer_id": customer["id"]},
    ).json()
    project = client.post(
        "/assessment-projects", json={"name": "2026 Review", "customer_id": customer["id"]}
    ).json()

    return project, vendor, product, edition, module, environment


def _run_simulation(client, project, vendor, product, edition, module, environment):
    resp = client.post(
        "/analysis/simulation",
        json={
            "assessment_project_id": project["id"],
            "scenario_type": "add_product",
            "name": "Add Acme Shield",
            "vendor_id": vendor["id"],
            "product_id": product["id"],
            "edition_id": edition["id"],
            "environment_id": environment["id"],
            "module_ids": [module["id"]],
            "deployment_model": "Agent",
        },
    )
    assert resp.status_code == 200
    return resp.json()


def test_post_simulation_endpoint(client):
    project, vendor, product, edition, module, environment = _build_scenario(client)
    report = _run_simulation(client, project, vendor, product, edition, module, environment)

    assert report["assessment_project_id"] == project["id"]
    assert report["scenario_type"] == "add_product"
    assert report["coverage_delta"]["classification"] == "Improvement"
    assert report["proposed_coverage"]["overall_coverage_percentage"] == 100.0
    assert report["current_coverage"]["overall_coverage_percentage"] == 0.0
    assert isinstance(report["executive_summary"], list) and report["executive_summary"]

    # The real assessment must remain empty -- the simulated assignment was
    # never committed.
    dashboard_resp = client.get(f"/assessment-projects/{project['id']}/dashboard")
    assert dashboard_resp.status_code == 200
    assert dashboard_resp.json()["total_deployed_products"] == 0


def test_get_simulation_report_by_id(client):
    project, vendor, product, edition, module, environment = _build_scenario(client)
    report = _run_simulation(client, project, vendor, product, edition, module, environment)

    resp = client.get(f"/analysis/simulation/{report['id']}")
    assert resp.status_code == 200
    assert resp.json()["id"] == report["id"]
    assert resp.json()["name"] == "Add Acme Shield"


def test_get_simulation_returns_404_for_unknown_id(client):
    resp = client.get("/analysis/simulation/999999")
    assert resp.status_code == 404


def test_simulation_summary_endpoint(client):
    project, vendor, product, edition, module, environment = _build_scenario(client)
    report = _run_simulation(client, project, vendor, product, edition, module, environment)

    resp = client.get(
        "/analysis/simulation/summary", params={"simulation_id": report["id"]}
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["coverage_delta"]["classification"] == "Improvement"
    assert "current_coverage" not in body
    assert "proposed_overlap" not in body
    assert "capability_comparison" not in body


def test_simulation_export_json(client):
    project, vendor, product, edition, module, environment = _build_scenario(client)
    report = _run_simulation(client, project, vendor, product, edition, module, environment)

    resp = client.get(
        "/analysis/simulation/export",
        params={"simulation_id": report["id"], "format": "json"},
    )
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/json"
    assert "attachment" in resp.headers["content-disposition"]
    assert resp.json()["id"] == report["id"]


def test_simulation_export_excel(client):
    project, vendor, product, edition, module, environment = _build_scenario(client)
    report = _run_simulation(client, project, vendor, product, edition, module, environment)

    resp = client.get(
        "/analysis/simulation/export",
        params={"simulation_id": report["id"], "format": "excel"},
    )
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]

    workbook = load_workbook(BytesIO(resp.content))
    assert "Summary" in workbook.sheetnames
    assert "Deltas" in workbook.sheetnames
    assert "Executive Summary" in workbook.sheetnames
    assert "Capability Comparison" in workbook.sheetnames
    assert "Vendor Comparison" in workbook.sheetnames
    assert "Framework Comparison" in workbook.sheetnames


def test_simulation_export_pdf(client):
    project, vendor, product, edition, module, environment = _build_scenario(client)
    report = _run_simulation(client, project, vendor, product, edition, module, environment)

    resp = client.get(
        "/analysis/simulation/export",
        params={"simulation_id": report["id"], "format": "pdf"},
    )
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/pdf"
    assert resp.content[:5] == b"%PDF-"


def test_simulation_export_rejects_invalid_format(client):
    project, vendor, product, edition, module, environment = _build_scenario(client)
    report = _run_simulation(client, project, vendor, product, edition, module, environment)

    resp = client.get(
        "/analysis/simulation/export",
        params={"simulation_id": report["id"], "format": "csv"},
    )
    assert resp.status_code == 422


def test_simulation_rejects_missing_scenario_fields(client):
    project, _, _, _, _, environment = _build_scenario(client)

    resp = client.post(
        "/analysis/simulation",
        json={
            "assessment_project_id": project["id"],
            "scenario_type": "add_product",
            "environment_id": environment["id"],
        },
    )
    assert resp.status_code == 422


def test_simulation_routes_not_shadowed_by_dynamic_id_route(client):
    """Regression guard: /simulation/export and /simulation/summary must be
    matched as static routes, not swallowed by /simulation/{simulation_id}."""
    project, vendor, product, edition, module, environment = _build_scenario(client)
    report = _run_simulation(client, project, vendor, product, edition, module, environment)

    export_resp = client.get(
        "/analysis/simulation/export",
        params={"simulation_id": report["id"], "format": "json"},
    )
    summary_resp = client.get(
        "/analysis/simulation/summary", params={"simulation_id": report["id"]}
    )

    assert export_resp.status_code == 200
    assert summary_resp.status_code == 200
