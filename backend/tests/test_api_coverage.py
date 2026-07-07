from openpyxl import load_workbook
from io import BytesIO


def _build_scenario(client):
    domain = client.post("/domains", json={"name": "Endpoint Security"}).json()
    capability = client.post(
        "/capabilities", json={"name": "EDR", "code": "EDR-001", "domain_id": domain["id"]}
    ).json()
    vendor = client.post("/vendors", json={"name": "Acme"}).json()
    product = client.post(
        "/products", json={"name": "Shield", "vendor_id": vendor["id"]}
    ).json()
    edition = client.post(
        "/editions", json={"name": "Pro", "product_id": product["id"]}
    ).json()
    module = client.post(
        "/modules",
        json={"name": "Detector", "edition_id": edition["id"], "capability_ids": [capability["id"]]},
    ).json()
    customer = client.post("/customers", json={"name": "Acme Corp"}).json()
    environment = client.post(
        "/environments",
        json={"name": "Production", "environment_type": "Production", "customer_id": customer["id"]},
    ).json()
    project = client.post(
        "/assessment-projects", json={"name": "2026 Review", "customer_id": customer["id"]}
    ).json()
    client.post(
        "/product-assignments",
        json={
            "assessment_project_id": project["id"],
            "vendor_id": vendor["id"],
            "product_id": product["id"],
            "edition_id": edition["id"],
            "environment_id": environment["id"],
            "module_ids": [module["id"]],
            "deployment_model": "Agent",
            "deployment_status": "Deployed",
        },
    )
    return project


def test_get_coverage_endpoint(client):
    project = _build_scenario(client)

    resp = client.get(f"/analysis/coverage/{project['id']}")
    assert resp.status_code == 200
    body = resp.json()
    assert body["assessment_project_id"] == project["id"]
    assert body["total_capabilities"] == 1
    assert body["covered_capability_count"] == 1
    assert body["overall_coverage_percentage"] == 100.0


def test_post_coverage_endpoint(client):
    project = _build_scenario(client)

    resp = client.post("/analysis/coverage", json={"assessment_project_id": project["id"]})
    assert resp.status_code == 200
    assert resp.json()["covered_capability_count"] == 1


def test_coverage_returns_404_for_unknown_assessment(client):
    resp = client.get("/analysis/coverage/999999")
    assert resp.status_code == 404


def test_domain_summary_endpoint(client):
    project = _build_scenario(client)

    resp = client.get("/analysis/domain-summary", params={"assessment_id": project["id"]})
    assert resp.status_code == 200
    body = resp.json()
    assert len(body) == 1
    assert body[0]["domain_name"] == "Endpoint Security"
    assert body[0]["covered_count"] == 1
    assert body[0]["total_count"] == 1


def test_capability_matrix_endpoint(client):
    project = _build_scenario(client)

    resp = client.get("/analysis/capabilities", params={"assessment_id": project["id"]})
    assert resp.status_code == 200
    body = resp.json()
    assert len(body["covered"]) == 1
    assert len(body["missing"]) == 0
    assert len(body["duplicate"]) == 0


def test_coverage_export_json(client):
    project = _build_scenario(client)

    resp = client.get(f"/analysis/coverage/{project['id']}/export", params={"format": "json"})
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/json"
    assert "attachment" in resp.headers["content-disposition"]
    assert resp.json()["assessment_project_id"] == project["id"]


def test_coverage_export_excel(client):
    project = _build_scenario(client)

    resp = client.get(f"/analysis/coverage/{project['id']}/export", params={"format": "excel"})
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    assert ".xlsx" in resp.headers["content-disposition"]

    workbook = load_workbook(BytesIO(resp.content))
    assert "Summary" in workbook.sheetnames
    assert "Domain Coverage" in workbook.sheetnames
    assert "Covered Capabilities" in workbook.sheetnames
    assert "Missing Capabilities" in workbook.sheetnames
    assert "Duplicate Capabilities" in workbook.sheetnames


def test_coverage_export_pdf(client):
    project = _build_scenario(client)

    resp = client.get(f"/analysis/coverage/{project['id']}/export", params={"format": "pdf"})
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/pdf"
    assert resp.content[:5] == b"%PDF-"


def test_coverage_export_rejects_invalid_format(client):
    project = _build_scenario(client)

    resp = client.get(f"/analysis/coverage/{project['id']}/export", params={"format": "csv"})
    assert resp.status_code == 422
