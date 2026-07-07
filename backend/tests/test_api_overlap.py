from io import BytesIO

from openpyxl import load_workbook


def _build_scenario(client):
    domain = client.post("/domains", json={"name": "Endpoint Security"}).json()
    capability = client.post(
        "/capabilities", json={"name": "EDR", "code": "EDR-001", "domain_id": domain["id"]}
    ).json()

    vendor_a = client.post("/vendors", json={"name": "Acme"}).json()
    product_a = client.post(
        "/products", json={"name": "Shield", "vendor_id": vendor_a["id"]}
    ).json()
    edition_a = client.post(
        "/editions", json={"name": "Pro", "product_id": product_a["id"]}
    ).json()
    module_a = client.post(
        "/modules",
        json={"name": "Detector", "edition_id": edition_a["id"], "capability_ids": [capability["id"]]},
    ).json()

    vendor_b = client.post("/vendors", json={"name": "Globex"}).json()
    product_b = client.post(
        "/products", json={"name": "Guard", "vendor_id": vendor_b["id"]}
    ).json()
    edition_b = client.post(
        "/editions", json={"name": "Standard", "product_id": product_b["id"]}
    ).json()
    module_b = client.post(
        "/modules",
        json={"name": "Sentinel", "edition_id": edition_b["id"], "capability_ids": [capability["id"]]},
    ).json()

    customer = client.post("/customers", json={"name": "Acme Corp"}).json()
    environment = client.post(
        "/environments",
        json={"name": "Production", "environment_type": "Production", "customer_id": customer["id"]},
    ).json()
    project = client.post(
        "/assessment-projects", json={"name": "2026 Review", "customer_id": customer["id"]}
    ).json()

    client.post(
        "/product-assignments",
        json={
            "assessment_project_id": project["id"],
            "vendor_id": vendor_a["id"],
            "product_id": product_a["id"],
            "edition_id": edition_a["id"],
            "environment_id": environment["id"],
            "module_ids": [module_a["id"]],
            "deployment_model": "Agent",
            "deployment_status": "Deployed",
        },
    )
    client.post(
        "/product-assignments",
        json={
            "assessment_project_id": project["id"],
            "vendor_id": vendor_b["id"],
            "product_id": product_b["id"],
            "edition_id": edition_b["id"],
            "environment_id": environment["id"],
            "module_ids": [module_b["id"]],
            "deployment_model": "Agent",
            "deployment_status": "Deployed",
        },
    )
    return project


def test_get_overlap_report_endpoint(client):
    project = _build_scenario(client)

    resp = client.get(f"/analysis/overlap/{project['id']}")
    assert resp.status_code == 200
    body = resp.json()
    assert body["assessment_project_id"] == project["id"]
    assert body["duplicate_capability_count"] == 1
    assert body["cross_vendor_duplicate_count"] == 1
    assert len(body["product_overlaps"]) == 1
    assert len(body["vendor_summary"]) == 2


def test_post_overlap_endpoint(client):
    project = _build_scenario(client)

    resp = client.post("/analysis/overlap", json={"assessment_project_id": project["id"]})
    assert resp.status_code == 200
    assert resp.json()["duplicate_capability_count"] == 1


def test_overlap_returns_404_for_unknown_assessment(client):
    resp = client.get("/analysis/overlap/999999")
    assert resp.status_code == 404


def test_overlap_summary_endpoint(client):
    project = _build_scenario(client)

    resp = client.get("/analysis/overlap/summary", params={"assessment_id": project["id"]})
    assert resp.status_code == 200
    body = resp.json()
    assert body["duplicate_capability_count"] == 1
    assert "duplicate_capabilities" not in body
    assert "vendor_summary" not in body
    assert "product_overlaps" not in body


def test_overlap_export_json(client):
    project = _build_scenario(client)

    resp = client.get(
        "/analysis/overlap/export", params={"assessment_id": project["id"], "format": "json"}
    )
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/json"
    assert "attachment" in resp.headers["content-disposition"]
    assert resp.json()["assessment_project_id"] == project["id"]


def test_overlap_export_excel(client):
    project = _build_scenario(client)

    resp = client.get(
        "/analysis/overlap/export", params={"assessment_id": project["id"], "format": "excel"}
    )
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    assert ".xlsx" in resp.headers["content-disposition"]

    workbook = load_workbook(BytesIO(resp.content))
    assert "Summary" in workbook.sheetnames
    assert "Vendor Comparison" in workbook.sheetnames
    assert "Duplicate Capabilities" in workbook.sheetnames
    assert "Product Overlap" in workbook.sheetnames
    assert "Redundant Licenses" in workbook.sheetnames
    assert "Unused Capabilities" in workbook.sheetnames


def test_overlap_export_pdf(client):
    project = _build_scenario(client)

    resp = client.get(
        "/analysis/overlap/export", params={"assessment_id": project["id"], "format": "pdf"}
    )
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/pdf"
    assert resp.content[:5] == b"%PDF-"


def test_overlap_export_rejects_invalid_format(client):
    project = _build_scenario(client)

    resp = client.get(
        "/analysis/overlap/export", params={"assessment_id": project["id"], "format": "csv"}
    )
    assert resp.status_code == 422


def test_overlap_routes_not_shadowed_by_dynamic_assessment_id_route(client):
    """Regression guard: /overlap/export and /overlap/summary must be
    matched as static routes, not swallowed by /overlap/{assessment_id}."""
    project = _build_scenario(client)

    export_resp = client.get(
        "/analysis/overlap/export", params={"assessment_id": project["id"], "format": "json"}
    )
    summary_resp = client.get(
        "/analysis/overlap/summary", params={"assessment_id": project["id"]}
    )

    assert export_resp.status_code == 200
    assert summary_resp.status_code == 200
