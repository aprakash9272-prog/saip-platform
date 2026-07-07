from io import BytesIO

from openpyxl import load_workbook


def _build_scenario(client):
    domain = client.post("/domains", json={"name": "Endpoint Security"}).json()
    covered_capability = client.post(
        "/capabilities", json={"name": "EDR", "code": "EDR-001", "domain_id": domain["id"]}
    ).json()
    missing_capability = client.post(
        "/capabilities",
        json={
            "name": "Threat Hunting",
            "code": "EDR-002",
            "domain_id": domain["id"],
            "risk_category": "Critical",
        },
    ).json()

    vendor = client.post("/vendors", json={"name": "Acme"}).json()
    product = client.post(
        "/products", json={"name": "Shield", "vendor_id": vendor["id"]}
    ).json()
    edition = client.post(
        "/editions", json={"name": "Pro", "product_id": product["id"]}
    ).json()
    module = client.post(
        "/modules",
        json={
            "name": "Detector",
            "edition_id": edition["id"],
            "capability_ids": [covered_capability["id"]],
        },
    ).json()
    customer = client.post("/customers", json={"name": "Acme Corp"}).json()
    environment = client.post(
        "/environments",
        json={"name": "Production", "environment_type": "Production", "customer_id": customer["id"]},
    ).json()
    project = client.post(
        "/assessment-projects", json={"name": "2026 Review", "customer_id": customer["id"]}
    ).json()
    client.post(
        "/product-assignments",
        json={
            "assessment_project_id": project["id"],
            "vendor_id": vendor["id"],
            "product_id": product["id"],
            "edition_id": edition["id"],
            "environment_id": environment["id"],
            "module_ids": [module["id"]],
            "deployment_model": "Agent",
            "deployment_status": "Deployed",
        },
    )
    return project, missing_capability


def test_get_gap_report_endpoint(client):
    project, missing_capability = _build_scenario(client)

    resp = client.get(f"/analysis/gaps/{project['id']}")
    assert resp.status_code == 200
    body = resp.json()
    assert body["assessment_project_id"] == project["id"]
    assert body["total_gaps"] == 1
    assert body["gaps"][0]["code"] == missing_capability["code"]
    assert body["gaps"][0]["status"] == "Open"


def test_post_gaps_endpoint(client):
    project, _ = _build_scenario(client)

    resp = client.post("/analysis/gaps", json={"assessment_project_id": project["id"]})
    assert resp.status_code == 200
    assert resp.json()["total_gaps"] == 1


def test_gaps_returns_404_for_unknown_assessment(client):
    resp = client.get("/analysis/gaps/999999")
    assert resp.status_code == 404


def test_gap_summary_endpoint(client):
    project, _ = _build_scenario(client)

    resp = client.get("/analysis/gaps/summary", params={"assessment_id": project["id"]})
    assert resp.status_code == 200
    body = resp.json()
    assert body["total_gaps"] == 1
    assert "gaps" not in body
    assert "domain_gap_scores" not in body


def test_gap_domains_endpoint(client):
    project, _ = _build_scenario(client)

    resp = client.get("/analysis/gaps/domains", params={"assessment_id": project["id"]})
    assert resp.status_code == 200
    body = resp.json()
    assert len(body) == 1
    assert body[0]["domain_name"] == "Endpoint Security"
    assert body[0]["missing_count"] == 1


def test_gap_export_json(client):
    project, _ = _build_scenario(client)

    resp = client.get("/analysis/gaps/export", params={"assessment_id": project["id"], "format": "json"})
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/json"
    assert "attachment" in resp.headers["content-disposition"]
    assert resp.json()["assessment_project_id"] == project["id"]


def test_gap_export_excel(client):
    project, _ = _build_scenario(client)

    resp = client.get("/analysis/gaps/export", params={"assessment_id": project["id"], "format": "excel"})
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    assert ".xlsx" in resp.headers["content-disposition"]

    workbook = load_workbook(BytesIO(resp.content))
    assert "Summary" in workbook.sheetnames
    assert "Domain Gap Scores" in workbook.sheetnames
    assert "Gaps" in workbook.sheetnames


def test_gap_export_pdf(client):
    project, _ = _build_scenario(client)

    resp = client.get("/analysis/gaps/export", params={"assessment_id": project["id"], "format": "pdf"})
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/pdf"
    assert resp.content[:5] == b"%PDF-"


def test_gap_export_rejects_invalid_format(client):
    project, _ = _build_scenario(client)

    resp = client.get("/analysis/gaps/export", params={"assessment_id": project["id"], "format": "csv"})
    assert resp.status_code == 422


def test_gap_export_route_not_shadowed_by_dynamic_assessment_id_route(client):
    """Regression guard: /gaps/export, /gaps/summary, /gaps/domains must be
    matched as static routes, not swallowed by /gaps/{assessment_id}."""
    project, _ = _build_scenario(client)

    export_resp = client.get(
        "/analysis/gaps/export", params={"assessment_id": project["id"], "format": "json"}
    )
    summary_resp = client.get("/analysis/gaps/summary", params={"assessment_id": project["id"]})
    domains_resp = client.get("/analysis/gaps/domains", params={"assessment_id": project["id"]})

    assert export_resp.status_code == 200
    assert summary_resp.status_code == 200
    assert domains_resp.status_code == 200
