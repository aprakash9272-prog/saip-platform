from io import BytesIO

from openpyxl import load_workbook


def _build_scenario(client):
    domain = client.post("/domains", json={"name": "Endpoint Security"}).json()
    gap_capability = client.post(
        "/capabilities",
        json={
            "name": "Threat Hunting",
            "code": "EDR-002",
            "domain_id": domain["id"],
            "risk_category": "Critical",
        },
    ).json()

    vendor = client.post("/vendors", json={"name": "Acme"}).json()
    product = client.post(
        "/products", json={"name": "Shield", "vendor_id": vendor["id"]}
    ).json()
    edition = client.post(
        "/editions", json={"name": "Pro", "product_id": product["id"]}
    ).json()
    module = client.post(
        "/modules",
        json={"name": "Hunter", "edition_id": edition["id"], "capability_ids": [gap_capability["id"]]},
    ).json()
    client.post(
        "/product-mappings",
        json={
            "vendor_id": vendor["id"],
            "product_id": product["id"],
            "edition_id": edition["id"],
            "module_id": module["id"],
            "capability_id": gap_capability["id"],
            "deployment_model": "SaaS",
        },
    )

    customer = client.post("/customers", json={"name": "Acme Corp"}).json()
    project = client.post(
        "/assessment-projects", json={"name": "2026 Review", "customer_id": customer["id"]}
    ).json()
    return project, gap_capability


def test_get_recommendation_report_endpoint(client):
    project, gap_capability = _build_scenario(client)

    resp = client.get(f"/analysis/recommendations/{project['id']}")
    assert resp.status_code == 200
    body = resp.json()
    assert body["assessment_project_id"] == project["id"]
    assert body["addressable_gaps"] == 1
    assert body["recommendations"][0]["capability_code"] == gap_capability["code"]
    assert body["recommendations"][0]["candidates"][0]["vendor"] == "Acme"


def test_post_recommendations_endpoint(client):
    project, _ = _build_scenario(client)

    resp = client.post(
        "/analysis/recommendations", json={"assessment_project_id": project["id"]}
    )
    assert resp.status_code == 200
    assert resp.json()["addressable_gaps"] == 1


def test_recommendations_returns_404_for_unknown_assessment(client):
    resp = client.get("/analysis/recommendations/999999")
    assert resp.status_code == 404


def test_recommendation_summary_endpoint(client):
    project, _ = _build_scenario(client)

    resp = client.get("/analysis/recommendations/summary", params={"assessment_id": project["id"]})
    assert resp.status_code == 200
    body = resp.json()
    assert body["addressable_gaps"] == 1
    assert "recommendations" not in body
    assert "priority_matrix" not in body
    assert "product_comparison" not in body


def test_recommendation_export_json(client):
    project, _ = _build_scenario(client)

    resp = client.get(
        "/analysis/recommendations/export", params={"assessment_id": project["id"], "format": "json"}
    )
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/json"
    assert "attachment" in resp.headers["content-disposition"]
    assert resp.json()["assessment_project_id"] == project["id"]


def test_recommendation_export_excel(client):
    project, _ = _build_scenario(client)

    resp = client.get(
        "/analysis/recommendations/export", params={"assessment_id": project["id"], "format": "excel"}
    )
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    assert ".xlsx" in resp.headers["content-disposition"]

    workbook = load_workbook(BytesIO(resp.content))
    assert "Summary" in workbook.sheetnames
    assert "Priority Matrix" in workbook.sheetnames
    assert "Product Comparison" in workbook.sheetnames
    assert "Recommendations" in workbook.sheetnames


def test_recommendation_export_pdf(client):
    project, _ = _build_scenario(client)

    resp = client.get(
        "/analysis/recommendations/export", params={"assessment_id": project["id"], "format": "pdf"}
    )
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/pdf"
    assert resp.content[:5] == b"%PDF-"


def test_recommendation_export_rejects_invalid_format(client):
    project, _ = _build_scenario(client)

    resp = client.get(
        "/analysis/recommendations/export", params={"assessment_id": project["id"], "format": "csv"}
    )
    assert resp.status_code == 422


def test_recommendation_routes_not_shadowed_by_dynamic_assessment_id_route(client):
    """Regression guard: /recommendations/export and /recommendations/summary
    must be matched as static routes, not swallowed by
    /recommendations/{assessment_id}."""
    project, _ = _build_scenario(client)

    export_resp = client.get(
        "/analysis/recommendations/export", params={"assessment_id": project["id"], "format": "json"}
    )
    summary_resp = client.get(
        "/analysis/recommendations/summary", params={"assessment_id": project["id"]}
    )

    assert export_resp.status_code == 200
    assert summary_resp.status_code == 200
