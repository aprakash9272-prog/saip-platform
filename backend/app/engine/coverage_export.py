"""Export builders for a CoverageReport: JSON, Excel, and PDF."""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from app.schemas.coverage import CoverageReport

HEADER_FONT = Font(bold=True)


def export_coverage_json(report: CoverageReport) -> bytes:
    return report.model_dump_json(indent=2).encode("utf-8")


def _autosize_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        length = max((len(str(cell.value)) for cell in column_cells if cell.value is not None), default=10)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 60)


def export_coverage_excel(report: CoverageReport) -> bytes:
    workbook = Workbook()

    summary = workbook.active
    summary.title = "Summary"
    summary.append(["Assessment Project", report.assessment_project_name])
    summary.append(["Generated At", report.generated_at.isoformat()])
    summary.append(["Total Capabilities", report.total_capabilities])
    summary.append(["Covered Capabilities", report.covered_capability_count])
    summary.append(["Missing Capabilities", report.missing_capability_count])
    summary.append(["Duplicate Capabilities", report.duplicate_capability_count])
    summary.append(["Overall Coverage %", report.overall_coverage_percentage])
    for row in summary["A1:A7"]:
        row[0].font = HEADER_FONT
    _autosize_columns(summary)

    domain_sheet = workbook.create_sheet("Domain Coverage")
    domain_sheet.append(["Domain", "Covered", "Total", "Coverage %"])
    for cell in domain_sheet[1]:
        cell.font = HEADER_FONT
    for d in report.domain_coverage:
        domain_sheet.append([d.domain_name, d.covered_count, d.total_count, d.coverage_percentage])
    _autosize_columns(domain_sheet)

    def _capability_sheet(name: str, items) -> None:
        sheet = workbook.create_sheet(name)
        sheet.append(["Code", "Name", "Domain", "Provider Count", "Providers"])
        for cell in sheet[1]:
            cell.font = HEADER_FONT
        for item in items:
            sheet.append(
                [item.code, item.name, item.domain_name, item.provider_count, ", ".join(item.providers)]
            )
        _autosize_columns(sheet)

    _capability_sheet("Covered Capabilities", report.covered_capabilities)
    _capability_sheet("Missing Capabilities", report.missing_capabilities)
    _capability_sheet("Duplicate Capabilities", report.duplicate_capabilities)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def export_coverage_pdf(report: CoverageReport) -> bytes:
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=LETTER, title="Coverage Report")
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph(f"Coverage Report — {report.assessment_project_name}", styles["Title"]))
    story.append(Paragraph(f"Generated: {report.generated_at.isoformat()}", styles["Normal"]))
    story.append(Spacer(1, 0.25 * inch))

    summary_data = [
        ["Total Capabilities", str(report.total_capabilities)],
        ["Covered Capabilities", str(report.covered_capability_count)],
        ["Missing Capabilities", str(report.missing_capability_count)],
        ["Duplicate Capabilities", str(report.duplicate_capability_count)],
        ["Overall Coverage %", f"{report.overall_coverage_percentage}%"],
    ]
    summary_table = Table(summary_data, colWidths=[2.5 * inch, 2 * inch])
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), colors.whitesmoke),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.append(summary_table)
    story.append(Spacer(1, 0.3 * inch))

    story.append(Paragraph("Domain Coverage", styles["Heading2"]))
    domain_rows = [["Domain", "Covered", "Total", "Coverage %"]]
    domain_rows += [
        [d.domain_name, str(d.covered_count), str(d.total_count), f"{d.coverage_percentage}%"]
        for d in report.domain_coverage
    ]
    domain_table = Table(domain_rows, colWidths=[3 * inch, 0.8 * inch, 0.8 * inch, 1 * inch])
    domain_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.whitesmoke]),
            ]
        )
    )
    story.append(domain_table)
    story.append(Spacer(1, 0.3 * inch))

    def _capability_section(title: str, items) -> None:
        story.append(Paragraph(f"{title} ({len(items)})", styles["Heading2"]))
        if not items:
            story.append(Paragraph("None.", styles["Normal"]))
            story.append(Spacer(1, 0.2 * inch))
            return
        rows = [["Code", "Name", "Domain"]]
        rows += [[item.code, item.name, item.domain_name] for item in items]
        table = Table(rows, colWidths=[1 * inch, 2.8 * inch, 2 * inch], repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.whitesmoke]),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                ]
            )
        )
        story.append(table)
        story.append(Spacer(1, 0.3 * inch))

    _capability_section("Missing Capabilities", report.missing_capabilities)
    _capability_section("Duplicate Capabilities", report.duplicate_capabilities)
    _capability_section("Covered Capabilities", report.covered_capabilities)

    doc.build(story)
    return buffer.getvalue()
