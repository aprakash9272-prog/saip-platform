"""Export builders for a GapReport: JSON, Excel, and PDF."""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from app.schemas.gap import SEVERITY_LEVELS, GapReport

HEADER_FONT = Font(bold=True)

_SEVERITY_COLORS = {
    "Critical": colors.HexColor("#7f1d1d"),
    "High": colors.HexColor("#b91c1c"),
    "Medium": colors.HexColor("#b45309"),
    "Low": colors.HexColor("#a16207"),
    "Informational": colors.HexColor("#6b7280"),
}


def export_gap_json(report: GapReport) -> bytes:
    return report.model_dump_json(indent=2).encode("utf-8")


def _autosize_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        length = max(
            (len(str(cell.value)) for cell in column_cells if cell.value is not None),
            default=10,
        )
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 60)


def export_gap_excel(report: GapReport) -> bytes:
    workbook = Workbook()

    summary = workbook.active
    summary.title = "Summary"
    summary.append(["Assessment Project", report.assessment_project_name])
    summary.append(["Generated At", report.generated_at.isoformat()])
    summary.append(["Total Capabilities", report.total_capabilities])
    summary.append(["Total Gaps", report.total_gaps])
    summary.append(["Critical", report.critical_count])
    summary.append(["High", report.high_count])
    summary.append(["Medium", report.medium_count])
    summary.append(["Low", report.low_count])
    summary.append(["Informational", report.informational_count])
    summary.append(["Overall Gap %", report.overall_gap_percentage])
    summary.append(["Overall Risk Score", report.overall_risk_score])
    for row in summary["A1:A11"]:
        row[0].font = HEADER_FONT
    _autosize_columns(summary)

    domain_sheet = workbook.create_sheet("Domain Gap Scores")
    domain_sheet.append(
        ["Domain", "Coverage %", "Gap %", "Missing Count", "Critical Gaps", "Risk Score"]
    )
    for cell in domain_sheet[1]:
        cell.font = HEADER_FONT
    for d in report.domain_gap_scores:
        domain_sheet.append(
            [
                d.domain_name,
                d.coverage_percentage,
                d.gap_percentage,
                d.missing_count,
                d.critical_gap_count,
                d.domain_risk_score,
            ]
        )
    _autosize_columns(domain_sheet)

    gaps_sheet = workbook.create_sheet("Gaps")
    gaps_sheet.append(
        [
            "Code",
            "Name",
            "Domain",
            "Risk Category",
            "Severity",
            "Business Impact",
            "Framework Controls",
            "Mapped Products",
            "Status",
        ]
    )
    for cell in gaps_sheet[1]:
        cell.font = HEADER_FONT
    for gap in sorted(report.gaps, key=lambda g: (SEVERITY_LEVELS.index(g.severity), g.code)):
        controls = "; ".join(
            f"{c.framework_name} {c.framework_version}: {c.control_id}"
            for c in gap.framework_controls
        )
        gaps_sheet.append(
            [
                gap.code,
                gap.name,
                gap.domain_name,
                gap.risk_category or "—",
                gap.severity,
                gap.business_impact,
                controls or "—",
                ", ".join(gap.mapped_products) or "—",
                gap.status,
            ]
        )
    _autosize_columns(gaps_sheet)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def export_gap_pdf(report: GapReport) -> bytes:
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=LETTER, title="Gap Analysis Report")
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph(f"Gap Analysis Report — {report.assessment_project_name}", styles["Title"]))
    story.append(Paragraph(f"Generated: {report.generated_at.isoformat()}", styles["Normal"]))
    story.append(Spacer(1, 0.25 * inch))

    summary_data = [
        ["Total Capabilities", str(report.total_capabilities)],
        ["Total Gaps", str(report.total_gaps)],
        ["Critical / High / Medium / Low / Informational",
         f"{report.critical_count} / {report.high_count} / {report.medium_count} / "
         f"{report.low_count} / {report.informational_count}"],
        ["Overall Gap %", f"{report.overall_gap_percentage}%"],
        ["Overall Risk Score", str(report.overall_risk_score)],
    ]
    summary_table = Table(summary_data, colWidths=[3 * inch, 3 * inch])
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), colors.whitesmoke),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
            ]
        )
    )
    story.append(summary_table)
    story.append(Spacer(1, 0.3 * inch))

    story.append(Paragraph("Domain Gap Scores", styles["Heading2"]))
    domain_rows = [["Domain", "Coverage %", "Gap %", "Missing", "Critical", "Risk Score"]]
    domain_rows += [
        [
            d.domain_name,
            f"{d.coverage_percentage}%",
            f"{d.gap_percentage}%",
            str(d.missing_count),
            str(d.critical_gap_count),
            str(d.domain_risk_score),
        ]
        for d in report.domain_gap_scores
    ]
    domain_table = Table(
        domain_rows,
        colWidths=[2.2 * inch, 0.8 * inch, 0.7 * inch, 0.7 * inch, 0.6 * inch, 0.8 * inch],
        repeatRows=1,
    )
    domain_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.whitesmoke]),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
            ]
        )
    )
    story.append(domain_table)
    story.append(Spacer(1, 0.3 * inch))

    story.append(Paragraph(f"Gaps ({report.total_gaps})", styles["Heading2"]))
    if not report.gaps:
        story.append(Paragraph("No gaps identified.", styles["Normal"]))
    else:
        gap_rows = [["Code", "Name", "Domain", "Severity", "Business Impact"]]
        sorted_gaps = sorted(
            report.gaps, key=lambda g: (SEVERITY_LEVELS.index(g.severity), g.code)
        )
        gap_rows += [
            [g.code, g.name, g.domain_name, g.severity, g.business_impact]
            for g in sorted_gaps
        ]
        gap_table = Table(
            gap_rows,
            colWidths=[0.9 * inch, 1.8 * inch, 1.6 * inch, 0.9 * inch, 1.1 * inch],
            repeatRows=1,
        )
        style_commands = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.whitesmoke]),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ]
        for row_index, gap in enumerate(sorted_gaps, start=1):
            style_commands.append(
                ("TEXTCOLOR", (3, row_index), (3, row_index), _SEVERITY_COLORS[gap.severity])
            )
            style_commands.append(("FONTNAME", (3, row_index), (3, row_index), "Helvetica-Bold"))
        gap_table.setStyle(TableStyle(style_commands))
        story.append(gap_table)

    doc.build(story)
    return buffer.getvalue()
