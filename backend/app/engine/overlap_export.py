"""Export builders for an OverlapReport: JSON, Excel, and PDF."""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from app.schemas.overlap import OverlapReport

HEADER_FONT = Font(bold=True)


def export_overlap_json(report: OverlapReport) -> bytes:
    return report.model_dump_json(indent=2).encode("utf-8")


def _autosize_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        length = max(
            (len(str(cell.value)) for cell in column_cells if cell.value is not None),
            default=10,
        )
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 60)


def export_overlap_excel(report: OverlapReport) -> bytes:
    workbook = Workbook()

    summary = workbook.active
    summary.title = "Summary"
    summary.append(["Assessment Project", report.assessment_project_name])
    summary.append(["Generated At", report.generated_at.isoformat()])
    summary.append(["Total Deployed Products", report.total_deployed_products])
    summary.append(["Total Vendors", report.total_vendors])
    summary.append(["Duplicate Capabilities", report.duplicate_capability_count])
    summary.append(["Cross-Vendor Duplicates", report.cross_vendor_duplicate_count])
    summary.append(["Unused Capabilities", report.unused_capability_count])
    summary.append(["Overlap %", report.overlap_percentage])
    summary.append(["Optimization Score", report.optimization_score])
    summary.append(["Vendor Consolidation Score", report.vendor_consolidation_score])
    summary.append(["License Reduction Opportunity", report.license_reduction_opportunity])
    summary.append(["Cost Optimization Score", report.cost_optimization_score])
    summary.append(["Operational Complexity Score", report.operational_complexity_score])
    for row in summary["A1:A13"]:
        row[0].font = HEADER_FONT
    _autosize_columns(summary)

    vendor_sheet = workbook.create_sheet("Vendor Comparison")
    vendor_sheet.append(
        [
            "Vendor",
            "Deployed Products",
            "Total Capabilities",
            "Unique",
            "Overlapping",
            "License Qty",
            "Open Gaps Addressable",
        ]
    )
    for cell in vendor_sheet[1]:
        cell.font = HEADER_FONT
    for v in report.vendor_summary:
        vendor_sheet.append(
            [
                v.vendor,
                v.deployed_product_count,
                v.total_capabilities_provided,
                v.unique_capabilities_provided,
                v.overlapping_capabilities_provided,
                v.total_license_quantity,
                v.open_gaps_addressable,
            ]
        )
    _autosize_columns(vendor_sheet)

    dup_sheet = workbook.create_sheet("Duplicate Capabilities")
    dup_sheet.append(["Code", "Name", "Domain", "Providers", "Vendors", "Cross-Vendor"])
    for cell in dup_sheet[1]:
        cell.font = HEADER_FONT
    for d in report.duplicate_capabilities:
        dup_sheet.append(
            [d.code, d.name, d.domain_name, ", ".join(d.providers), d.distinct_vendor_count, d.cross_vendor]
        )
    _autosize_columns(dup_sheet)

    product_sheet = workbook.create_sheet("Product Overlap")
    product_sheet.append(["Vendor A", "Product A", "Vendor B", "Product B", "Shared Capabilities", "Overlap %"])
    for cell in product_sheet[1]:
        cell.font = HEADER_FONT
    for p in report.product_overlaps:
        product_sheet.append(
            [p.vendor_a, p.product_a, p.vendor_b, p.product_b, p.shared_capability_count, p.overlap_percentage]
        )
    _autosize_columns(product_sheet)

    license_sheet = workbook.create_sheet("Redundant Licenses")
    license_sheet.append(
        ["Vendor", "Product", "Edition", "License Qty", "Redundant Caps", "Total Caps", "Redundancy %", "Fully Redundant"]
    )
    for cell in license_sheet[1]:
        cell.font = HEADER_FONT
    for r in report.redundant_licenses:
        license_sheet.append(
            [
                r.vendor,
                r.product,
                r.edition,
                r.license_quantity if r.license_quantity is not None else "—",
                r.redundant_capability_count,
                r.total_capability_count,
                r.redundancy_percentage,
                r.fully_redundant,
            ]
        )
    _autosize_columns(license_sheet)

    unused_sheet = workbook.create_sheet("Unused Capabilities")
    unused_sheet.append(["Vendor", "Product", "Edition", "Module", "Capability", "Domain"])
    for cell in unused_sheet[1]:
        cell.font = HEADER_FONT
    for u in report.unused_capabilities:
        unused_sheet.append(
            [u.vendor, u.product, u.edition, u.module, f"{u.capability_code} — {u.capability_name}", u.domain_name]
        )
    _autosize_columns(unused_sheet)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def export_overlap_pdf(report: OverlapReport) -> bytes:
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=LETTER, title="Overlap & Optimization Report")
    styles = getSampleStyleSheet()
    story = []

    story.append(
        Paragraph(f"Overlap & Optimization Report — {report.assessment_project_name}", styles["Title"])
    )
    story.append(Paragraph(f"Generated: {report.generated_at.isoformat()}", styles["Normal"]))
    story.append(Spacer(1, 0.25 * inch))

    summary_data = [
        ["Deployed Products / Vendors", f"{report.total_deployed_products} / {report.total_vendors}"],
        ["Duplicate Capabilities (cross-vendor)",
         f"{report.duplicate_capability_count} ({report.cross_vendor_duplicate_count})"],
        ["Unused Capabilities", str(report.unused_capability_count)],
        ["Overlap %", f"{report.overlap_percentage}%"],
        ["Optimization Score", str(report.optimization_score)],
        ["Vendor Consolidation Score", str(report.vendor_consolidation_score)],
        ["License Reduction Opportunity", str(report.license_reduction_opportunity)],
        ["Cost Optimization Score", str(report.cost_optimization_score)],
        ["Operational Complexity Score", str(report.operational_complexity_score)],
    ]
    summary_table = Table(summary_data, colWidths=[3 * inch, 3 * inch])
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), colors.whitesmoke),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
            ]
        )
    )
    story.append(summary_table)
    story.append(Spacer(1, 0.3 * inch))

    story.append(Paragraph("Vendor Comparison", styles["Heading2"]))
    if report.vendor_summary:
        vendor_rows = [["Vendor", "Products", "Total Caps", "Unique", "Overlapping", "Open Gaps"]]
        vendor_rows += [
            [v.vendor, str(v.deployed_product_count), str(v.total_capabilities_provided),
             str(v.unique_capabilities_provided), str(v.overlapping_capabilities_provided),
             str(v.open_gaps_addressable)]
            for v in report.vendor_summary
        ]
        vendor_table = Table(
            vendor_rows,
            colWidths=[1.3 * inch, 0.8 * inch, 0.9 * inch, 0.8 * inch, 1 * inch, 0.9 * inch],
            repeatRows=1,
        )
        vendor_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.whitesmoke]),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                ]
            )
        )
        story.append(vendor_table)
    else:
        story.append(Paragraph("No deployed vendors.", styles["Normal"]))
    story.append(Spacer(1, 0.3 * inch))

    story.append(Paragraph(f"Duplicate Capabilities ({len(report.duplicate_capabilities)})", styles["Heading2"]))
    if report.duplicate_capabilities:
        dup_rows = [["Code", "Name", "Domain", "Providers", "Cross-Vendor"]]
        dup_rows += [
            [d.code, d.name, d.domain_name, str(d.provider_count), "Yes" if d.cross_vendor else "No"]
            for d in report.duplicate_capabilities
        ]
        dup_table = Table(
            dup_rows, colWidths=[0.9 * inch, 1.8 * inch, 1.6 * inch, 0.9 * inch, 0.9 * inch], repeatRows=1
        )
        dup_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.whitesmoke]),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                ]
            )
        )
        story.append(dup_table)
    else:
        story.append(Paragraph("No duplicate capabilities.", styles["Normal"]))

    doc.build(story)
    return buffer.getvalue()
