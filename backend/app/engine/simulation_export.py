"""Export builders for a SimulationReport: JSON, Excel, and PDF."""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from app.schemas.simulation import SimulationReport

HEADER_FONT = Font(bold=True)

_DELTA_FIELDS = (
    "coverage_delta",
    "gap_delta",
    "overlap_delta",
    "recommendation_delta",
    "risk_delta",
    "cost_delta",
    "complexity_delta",
    "vendor_count_delta",
    "license_count_delta",
    "framework_coverage_delta",
)


def export_simulation_json(report: SimulationReport) -> bytes:
    return report.model_dump_json(indent=2).encode("utf-8")


def _autosize_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        length = max(
            (len(str(cell.value)) for cell in column_cells if cell.value is not None),
            default=10,
        )
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 60)


def export_simulation_excel(report: SimulationReport) -> bytes:
    workbook = Workbook()

    summary = workbook.active
    summary.title = "Summary"
    summary.append(["Assessment Project", report.assessment_project_name])
    summary.append(["Scenario", report.name or report.scenario_type.value])
    summary.append(["Scenario Type", report.scenario_type.value])
    summary.append(["Generated At", report.generated_at.isoformat()])
    for row in summary["A1:A4"]:
        row[0].font = HEADER_FONT
    _autosize_columns(summary)

    delta_sheet = workbook.create_sheet("Deltas")
    delta_sheet.append(
        ["Metric", "Current", "Proposed", "Delta", "% Change", "Classification"]
    )
    for cell in delta_sheet[1]:
        cell.font = HEADER_FONT
    for field in _DELTA_FIELDS:
        delta = getattr(report, field)
        delta_sheet.append(
            [
                delta.metric,
                delta.current_value,
                delta.proposed_value,
                delta.delta,
                delta.percentage_change,
                delta.classification.value,
            ]
        )
    _autosize_columns(delta_sheet)

    summary_sheet = workbook.create_sheet("Executive Summary")
    for line in report.executive_summary:
        summary_sheet.append([line])
    _autosize_columns(summary_sheet)

    capability_sheet = workbook.create_sheet("Capability Comparison")
    capability_sheet.append(
        ["Code", "Name", "Domain", "Current Covered", "Proposed Covered", "Classification"]
    )
    for cell in capability_sheet[1]:
        cell.font = HEADER_FONT
    for c in report.capability_comparison:
        capability_sheet.append(
            [c.code, c.name, c.domain_name, c.current_covered, c.proposed_covered, c.classification.value]
        )
    _autosize_columns(capability_sheet)

    vendor_sheet = workbook.create_sheet("Vendor Comparison")
    vendor_sheet.append(
        [
            "Vendor",
            "Current Deployed",
            "Proposed Deployed",
            "Current Capabilities",
            "Proposed Capabilities",
            "Classification",
        ]
    )
    for cell in vendor_sheet[1]:
        cell.font = HEADER_FONT
    for v in report.vendor_comparison:
        vendor_sheet.append(
            [
                v.vendor,
                v.current_deployed,
                v.proposed_deployed,
                v.current_capability_count,
                v.proposed_capability_count,
                v.classification.value,
            ]
        )
    _autosize_columns(vendor_sheet)

    framework_sheet = workbook.create_sheet("Framework Comparison")
    framework_sheet.append(
        ["Framework", "Version", "Total Controls", "Current Satisfied", "Proposed Satisfied", "Classification"]
    )
    for cell in framework_sheet[1]:
        cell.font = HEADER_FONT
    for f in report.framework_comparison:
        framework_sheet.append(
            [
                f.framework_name,
                f.framework_version,
                f.total_controls,
                f.current_satisfied_controls,
                f.proposed_satisfied_controls,
                f.classification.value,
            ]
        )
    _autosize_columns(framework_sheet)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def export_simulation_pdf(report: SimulationReport) -> bytes:
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=LETTER, title="Scenario Simulation Report")
    styles = getSampleStyleSheet()
    story = []

    title = report.name or report.scenario_type.value.replace("_", " ").title()
    story.append(
        Paragraph(f"Scenario Simulation Report — {title}", styles["Title"])
    )
    story.append(
        Paragraph(f"Assessment Project: {report.assessment_project_name}", styles["Normal"])
    )
    story.append(Paragraph(f"Generated: {report.generated_at.isoformat()}", styles["Normal"]))
    story.append(Spacer(1, 0.25 * inch))

    story.append(Paragraph("Executive Summary", styles["Heading2"]))
    for line in report.executive_summary:
        story.append(Paragraph(line, styles["Normal"]))
    story.append(Spacer(1, 0.3 * inch))

    story.append(Paragraph("Current vs Proposed Deltas", styles["Heading2"]))
    delta_rows = [["Metric", "Current", "Proposed", "Delta", "% Change", "Classification"]]
    for field in _DELTA_FIELDS:
        delta = getattr(report, field)
        delta_rows.append(
            [
                delta.metric,
                str(delta.current_value),
                str(delta.proposed_value),
                str(delta.delta),
                f"{delta.percentage_change:+.2f}%",
                delta.classification.value,
            ]
        )
    delta_table = Table(
        delta_rows,
        colWidths=[2.0 * inch, 0.8 * inch, 0.8 * inch, 0.7 * inch, 0.8 * inch, 0.9 * inch],
        repeatRows=1,
    )
    delta_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.whitesmoke]),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
            ]
        )
    )
    story.append(delta_table)
    story.append(Spacer(1, 0.3 * inch))

    story.append(
        Paragraph(f"Capability Changes ({len(report.capability_comparison)})", styles["Heading2"])
    )
    if report.capability_comparison:
        cap_rows = [["Code", "Name", "Domain", "Current", "Proposed", "Classification"]]
        cap_rows += [
            [
                c.code,
                c.name,
                c.domain_name,
                "Covered" if c.current_covered else "Missing",
                "Covered" if c.proposed_covered else "Missing",
                c.classification.value,
            ]
            for c in report.capability_comparison[:200]
        ]
        cap_table = Table(
            cap_rows,
            colWidths=[0.8 * inch, 1.6 * inch, 1.3 * inch, 0.8 * inch, 0.8 * inch, 0.9 * inch],
            repeatRows=1,
        )
        cap_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.whitesmoke]),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                ]
            )
        )
        story.append(cap_table)
    else:
        story.append(Paragraph("No capability status changes.", styles["Normal"]))

    doc.build(story)
    return buffer.getvalue()
