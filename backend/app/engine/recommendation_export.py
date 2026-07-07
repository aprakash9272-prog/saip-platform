"""Export builders for a RecommendationReport: JSON, Excel, and PDF."""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from app.schemas.recommendation import PRIORITY_LEVELS, RecommendationReport

HEADER_FONT = Font(bold=True)


def export_recommendation_json(report: RecommendationReport) -> bytes:
    return report.model_dump_json(indent=2).encode("utf-8")


def _autosize_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        length = max(
            (len(str(cell.value)) for cell in column_cells if cell.value is not None),
            default=10,
        )
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 60)


def export_recommendation_excel(report: RecommendationReport) -> bytes:
    workbook = Workbook()

    summary = workbook.active
    summary.title = "Summary"
    summary.append(["Assessment Project", report.assessment_project_name])
    summary.append(["Generated At", report.generated_at.isoformat()])
    summary.append(["Total Gaps", report.total_gaps])
    summary.append(["Addressable Gaps", report.addressable_gaps])
    summary.append(["Unaddressable Gaps", report.unaddressable_gaps])
    summary.append(["Critical Priority", report.critical_priority_count])
    summary.append(["High Priority", report.high_priority_count])
    summary.append(["Medium Priority", report.medium_priority_count])
    summary.append(["Low Priority", report.low_priority_count])
    summary.append(["Current Risk Score", report.current_risk_score])
    summary.append(["Projected Risk Score", report.projected_risk_score])
    summary.append(["Estimated Risk Reduction", report.estimated_overall_risk_reduction])
    summary.append(["Current Coverage %", report.coverage_forecast.current_coverage_percentage])
    summary.append(
        ["Projected Coverage %", report.coverage_forecast.projected_coverage_percentage]
    )
    for row in summary["A1:A14"]:
        row[0].font = HEADER_FONT
    _autosize_columns(summary)

    priority_sheet = workbook.create_sheet("Priority Matrix")
    priority_sheet.append(["Priority", "Count", "Capability Codes"])
    for cell in priority_sheet[1]:
        cell.font = HEADER_FONT
    for entry in report.priority_matrix:
        priority_sheet.append([entry.priority, entry.count, ", ".join(entry.capability_codes)])
    _autosize_columns(priority_sheet)

    comparison_sheet = workbook.create_sheet("Product Comparison")
    comparison_sheet.append(["Vendor", "Product", "Gaps Addressed", "Avg Confidence", "Domains Covered"])
    for cell in comparison_sheet[1]:
        cell.font = HEADER_FONT
    for entry in report.product_comparison:
        comparison_sheet.append(
            [
                entry.vendor,
                entry.product,
                entry.gaps_addressed,
                entry.average_confidence_score,
                ", ".join(entry.domains_covered),
            ]
        )
    _autosize_columns(comparison_sheet)

    rec_sheet = workbook.create_sheet("Recommendations")
    rec_sheet.append(
        [
            "Capability",
            "Domain",
            "Severity",
            "Priority",
            "Best Vendor",
            "Best Product",
            "Module",
            "License Tier",
            "Deployment Model",
            "Confidence",
            "Complexity",
            "Estimated Effort",
            "Risk Reduction",
        ]
    )
    for cell in rec_sheet[1]:
        cell.font = HEADER_FONT
    for rec in sorted(
        report.recommendations,
        key=lambda r: (PRIORITY_LEVELS.index(r.priority), r.capability_code),
    ):
        best = rec.candidates[0]
        rec_sheet.append(
            [
                f"{rec.capability_code} — {rec.capability_name}",
                rec.domain_name,
                rec.severity,
                rec.priority,
                best.vendor,
                best.product,
                best.module,
                best.licensing_tier or "—",
                best.deployment_model,
                best.confidence_score,
                best.implementation_complexity,
                best.estimated_effort,
                rec.estimated_risk_reduction,
            ]
        )
    _autosize_columns(rec_sheet)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def export_recommendation_pdf(report: RecommendationReport) -> bytes:
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=LETTER, title="Recommendation Report")
    styles = getSampleStyleSheet()
    story = []

    story.append(
        Paragraph(f"Recommendation Report — {report.assessment_project_name}", styles["Title"])
    )
    story.append(Paragraph(f"Generated: {report.generated_at.isoformat()}", styles["Normal"]))
    story.append(Spacer(1, 0.25 * inch))

    summary_data = [
        ["Total Gaps", str(report.total_gaps)],
        ["Addressable / Unaddressable", f"{report.addressable_gaps} / {report.unaddressable_gaps}"],
        [
            "Priority (Critical/High/Medium/Low)",
            f"{report.critical_priority_count} / {report.high_priority_count} / "
            f"{report.medium_priority_count} / {report.low_priority_count}",
        ],
        ["Current -> Projected Risk Score", f"{report.current_risk_score} -> {report.projected_risk_score}"],
        ["Estimated Risk Reduction", str(report.estimated_overall_risk_reduction)],
        [
            "Current -> Projected Coverage %",
            f"{report.coverage_forecast.current_coverage_percentage}% -> "
            f"{report.coverage_forecast.projected_coverage_percentage}%",
        ],
    ]
    summary_table = Table(summary_data, colWidths=[3 * inch, 3 * inch])
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), colors.whitesmoke),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
            ]
        )
    )
    story.append(summary_table)
    story.append(Spacer(1, 0.3 * inch))

    story.append(Paragraph("Product Comparison", styles["Heading2"]))
    if report.product_comparison:
        comparison_rows = [["Vendor", "Product", "Gaps Addressed", "Avg Confidence", "Domains"]]
        comparison_rows += [
            [
                entry.vendor,
                entry.product,
                str(entry.gaps_addressed),
                str(entry.average_confidence_score),
                ", ".join(entry.domains_covered),
            ]
            for entry in report.product_comparison[:15]
        ]
        comparison_table = Table(
            comparison_rows,
            colWidths=[1.3 * inch, 1.5 * inch, 0.9 * inch, 0.9 * inch, 1.8 * inch],
            repeatRows=1,
        )
        comparison_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.whitesmoke]),
                    ("FONTSIZE", (0, 0), (-1, -1), 7.5),
                ]
            )
        )
        story.append(comparison_table)
    else:
        story.append(Paragraph("No addressable gaps.", styles["Normal"]))
    story.append(Spacer(1, 0.3 * inch))

    story.append(Paragraph(f"Top Recommendations ({len(report.recommendations)})", styles["Heading2"]))
    if not report.recommendations:
        story.append(Paragraph("No recommendations — no catalog products match any gap.", styles["Normal"]))
    else:
        rec_rows = [["Capability", "Domain", "Priority", "Best Product", "Confidence", "Effort"]]
        sorted_recs = sorted(
            report.recommendations,
            key=lambda r: (PRIORITY_LEVELS.index(r.priority), r.capability_code),
        )
        for rec in sorted_recs:
            best = rec.candidates[0]
            rec_rows.append(
                [
                    rec.capability_code,
                    rec.domain_name,
                    rec.priority,
                    f"{best.vendor} - {best.product}",
                    str(best.confidence_score),
                    best.estimated_effort,
                ]
            )
        rec_table = Table(
            rec_rows,
            colWidths=[0.9 * inch, 1.5 * inch, 0.8 * inch, 1.8 * inch, 0.8 * inch, 1 * inch],
            repeatRows=1,
        )
        rec_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.whitesmoke]),
                    ("FONTSIZE", (0, 0), (-1, -1), 7.5),
                ]
            )
        )
        story.append(rec_table)

    doc.build(story)
    return buffer.getvalue()
